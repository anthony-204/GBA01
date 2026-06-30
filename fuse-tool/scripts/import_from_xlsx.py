#!/usr/bin/env python3
"""
Excel → JSON data migration for the Fuse Tool.

Reads Fuse_GUI_APP.xlsx and writes typed JSON files to fuse-tool/data/.
Run from repo root:  python scripts/import_from_xlsx.py

Column names are normalized (trimmed, line breaks removed) to fix Excel
inconsistent naming. See docs/DATA_MIGRATION.md for field mappings.
"""

from __future__ import annotations

import json
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Install openpyxl: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

ROOT = Path(__file__).resolve().parents[1]
XLSX_DEFAULT = ROOT.parent / "Resources" / "Tool" / "APP_Fuse" / "Fuse_GUI_APP.xlsx"
OUT_DIR = ROOT / "data"


def normalize_header(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).replace("\n", " ")).strip()


def normalize_key(header: str) -> str:
    """Convert header to camelCase id used in JSON."""
    h = normalize_header(header)
    mapping = {
        "#": "index",
        "Site": "site",
        "Manufcturer": "manufacturer",
        "Manufacturer": "manufacturer",
        "Model": "model",
        "TYPE": "type",
        "Category": "category",
        "Engine": "engine",
        "Electrical System (V)": "electricalSystemV",
        "Minimum battery Voltage (V)": "minBatteryVoltageV",
        "Configuration": "batteryConfiguration",
        "5 sec discharge": "fiveSecDischargeAs",
        "Fuse Installed (A)": "fuseInstalledA",
        "TyPe": "starterMotorType",
        "Cut off Voltage (V)": "cutoffVoltageV",
        "Power at Cut off voltage (kW)": "powerAtCutoffKw",
        "Efficiency (%)": "efficiencyPercent",
        "Peak current cut off(A) from power & eff calcukatuiion": "peakCurrentCutoffA",
        "Inrush Current (A) Measured": "inrushCurrentA",
        "Inrush current time (s)": "inrushTimeS",
        "Peak continuous current during cranking (A)": "peakCrankingCurrentA",
        "No of Motors": "motorCount",
        "type": "motorType",
        "Cranking Voltge (V) measured": "crankingVoltageMeasuredV",
        "Cranking time (s) Measured": "crankingTimeMeasuredS",
        "Cranking time (s)required": "crankingTimeRequiredS",
        "Alternator Continuous Current (A)": "alternatorContinuousA",
        "Type": "alternatorType",
        "Voltage (v)": "alternatorVoltageFactor",
        "No of Alternators": "alternatorCount",
        "Cable Type": "cableType",
        "Cable Size(mm2)": "cableSizeMm2",
        "Operating temperature (DegC)": "operatingTempC",
        "Continuous Current (A)": "cableContinuousA",
        "Peak (A)": "cablePeakA",
        "Cable Length (m)": "cableLengthM",
        "Link": "link",
    }
    if h in mapping:
        return mapping[h]
    # fallback: snake-ish
    key = re.sub(r"[^a-zA-Z0-9]+", "_", h).strip("_").lower()
    return key or "unknown"


def cell_value(v: object) -> object:
    if isinstance(v, str):
        s = v.strip()
        if s in ("", "TBC", "Select"):
            return None
        if s.lower() in ("data unavailable", "n/a", "#n/a"):
            return None
        return s
    return v


def import_machines(ws) -> list[dict]:
    headers: dict[int, str] = {}
    for c in range(1, ws.max_column + 1):
        h = ws.cell(2, c).value
        if h:
            headers[c] = normalize_key(str(h))

    machines: list[dict] = []
    for r in range(4, ws.max_row + 1):
        model = ws.cell(r, 4).value  # column D = Model
        if not model or str(model).strip() in ("Select", ""):
            continue
        record: dict = {"id": str(model).strip(), "sourceRow": r}
        for c, key in headers.items():
            raw = ws.cell(r, c).value
            val = cell_value(raw)
            if val is not None:
                record[key] = val
        machines.append(record)
    return machines


def import_fuse_library(ws) -> list[dict]:
    fuses: list[dict] = []
    for r in range(2, ws.max_row + 1):
        rating = ws.cell(r, 5).value
        if rating is None:
            continue
        fuses.append(
            {
                "gbPartHolder": cell_value(ws.cell(r, 1).value),
                "manufacturer": cell_value(ws.cell(r, 2).value),
                "description": normalize_header(ws.cell(r, 3).value),
                "manufacturerPartNumber": cell_value(ws.cell(r, 4).value),
                "currentRatingA": float(rating),
                "i2tA2s": cell_value(ws.cell(r, 6).value),
                "breakingCurrentA": cell_value(ws.cell(r, 7).value),
                "timeFromI2tS": cell_value(ws.cell(r, 8).value),
                "timeFromGraphS": cell_value(ws.cell(r, 9).value),
                "interruptingRating": cell_value(ws.cell(r, 10).value),
                "temperatureRangeC": normalize_header(ws.cell(r, 11).value),
                "link": cell_value(ws.cell(r, 12).value),
                "ratingOptionA": cell_value(ws.cell(r, 13).value),
            }
        )
    return fuses


def import_mega32v(ws) -> dict:
    ratings: list[float] = []
    for c in range(2, ws.max_column + 1):
        v = ws.cell(2, c).value
        if v is not None and isinstance(v, (int, float)):
            ratings.append(float(v))

    rows: list[dict] = []
    for r in range(3, ws.max_row + 1):
        crank = ws.cell(r, 1).value
        if crank is None:
            continue
        try:
            crank_a = float(crank)
        except (TypeError, ValueError):
            continue
        by_rating: dict[str, float | None] = {}
        for i, rating in enumerate(ratings):
            t = ws.cell(r, i + 2).value
            if t == "inf":
                by_rating[str(int(rating))] = None
            elif t is not None:
                try:
                    by_rating[str(int(rating))] = float(t)
                except (TypeError, ValueError):
                    pass
        rows.append({"crankingCurrentA": crank_a, "withstandTimeSByRating": by_rating})
    return {"fuseRatingsA": ratings, "rows": rows}


def import_cable_capacity(ws) -> list[dict]:
    cables: list[dict] = []
    for r in range(2, ws.max_row + 1):
        size = ws.cell(r, 6).value
        cap = ws.cell(r, 7).value
        if size is None or cap is None:
            continue
        try:
            cables.append(
                {
                    "cableType": normalize_header(ws.cell(r, 1).value),
                    "insulationType": normalize_header(ws.cell(r, 2).value),
                    "kFactor": float(ws.cell(r, 3).value) if ws.cell(r, 3).value else None,
                    "installationMethod": normalize_header(ws.cell(r, 4).value),
                    "maxConductorTempC": ws.cell(r, 5).value,
                    "sizeMm2": float(size),
                    "continuousCurrentA": float(cap),
                    "resistanceOhmPerKm": ws.cell(r, 8).value,
                }
            )
        except (TypeError, ValueError):
            continue
    return cables


def import_k_factors(ws) -> list[dict]:
    rows: list[dict] = []
    for r in range(3, ws.max_row + 1):
        ct = ws.cell(r, 1).value
        if not ct:
            continue
        rows.append(
            {
                "cableTypeLabel": normalize_header(ct),
                "kCopper": ws.cell(r, 2).value,
                "kCopperHighTemp": ws.cell(r, 3).value,
                "initialTempC": ws.cell(r, 6).value,
                "finalTempC": ws.cell(r, 7).value,
            }
        )
    return rows


def import_constants() -> dict:
    return {
        "defaultSafetyFactorPercent": 25,
        "defaultPeakCrankingLimitA": 1000,
        "minCrankingTimeRequiredS": 5,
        "maxAdiabaticCrankingTimeS": 5,
        "defaultKFactorCopper": 143,
        "defaultElectricalSystemV": 24,
        "minBatteryVoltage24V": 16.48,
        "voltageDropPercentLimit": 3,
        "standards": [
            "AS/NZS 5000.1 (design guidelines)",
            "AS/NZS 1125 (conductor)",
            "AS/NZS 1995:2003 (welding cable duty cycles — reference)",
        ],
        "notes": [
            "Constants align with MD6250 Project sheet and User Input defaults.",
            "peakCurrentCutoffA per machine overrides defaultPeakCrankingLimitA when present.",
        ],
    }


def main() -> None:
    xlsx = Path(sys.argv[1]) if len(sys.argv) > 1 else XLSX_DEFAULT
    if not xlsx.exists():
        print(f"Excel file not found: {xlsx}", file=sys.stderr)
        sys.exit(1)

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.load_workbook(xlsx, data_only=True)

    payload = {
        "meta": {
            "sourceFile": str(xlsx.name),
            "importedBy": "scripts/import_from_xlsx.py",
            "sheetNames": wb.sheetnames,
        },
        "constants": import_constants(),
        "machines": import_machines(wb["MachinesOnSite"]),
        "fuseLibrary": import_fuse_library(wb["Fuse_Library"]),
        "mega32vCurve": import_mega32v(wb["MEGA32V"]),
        "cableCapacity": import_cable_capacity(wb["Cable_Capacity"]),
        "copperKFactors": import_k_factors(wb["Copper_K_Factor"]),
    }

    for name, data in [
        ("constants.json", payload["constants"]),
        ("machines.json", payload["machines"]),
        ("fuse-library.json", payload["fuseLibrary"]),
        ("mega32v-curve.json", payload["mega32vCurve"]),
        ("cable-capacity.json", payload["cableCapacity"]),
        ("copper-k-factors.json", payload["copperKFactors"]),
        ("meta.json", payload["meta"]),
    ]:
        path = OUT_DIR / name
        with path.open("w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, default=str)
        print(f"Wrote {path}")

    # Combined bundle for web app static import
    bundle = OUT_DIR / "bundle.json"
    with bundle.open("w", encoding="utf-8") as f:
        json.dump(
            {
                "constants": payload["constants"],
                "machines": payload["machines"],
                "fuseLibrary": payload["fuseLibrary"],
                "mega32vCurve": payload["mega32vCurve"],
                "cableCapacity": payload["cableCapacity"],
                "copperKFactors": payload["copperKFactors"],
            },
            f,
            indent=2,
            default=str,
        )
    print(f"Wrote {bundle}")
    print(f"Imported {len(payload['machines'])} machines, {len(payload['fuseLibrary'])} fuses.")


if __name__ == "__main__":
    main()
