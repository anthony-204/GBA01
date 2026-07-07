#!/usr/bin/env python3
"""Export v0 machine calculation parameters to docs/V0_MACHINE_PARAMETERS.xlsx."""

from __future__ import annotations

import json
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Install openpyxl: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

ROOT = Path(__file__).resolve().parents[1]
DATA = ROOT / "data" / "bundle.json"
OUT = ROOT / "docs" / "V0_MACHINE_PARAMETERS.xlsx"

CLIENT_IDS = [
    "D10T",
    "B45E",
    "D11",
    "14M",
    "155 / D155AX-6",
    "375/ D375-5E0",
    "777 (07)",
    "793F",
    "992K",
]

K_MAP = {
    "OEM Wiring": 143,
    "Weldflex": 150,
    "weldflex": 150,
    "Narva": 150,
}

R_BY_SIZE = {50: 0.471, 70: 0.327, 95: 0.236, 120: 0.188}


def k_factor(cable_type: str, copper_rows: list) -> float | None:
    if not cable_type:
        return None
    low = cable_type.lower()
    for row in copper_rows:
        label = str(row.get("cableTypeLabel", "")).lower()
        if low in label or label in low:
            k = row.get("kCopper")
            if k is not None:
                return float(k)
    return K_MAP.get(cable_type)


def resistance(size: float | None, cable_rows: list) -> float | None:
    if size is None:
        return None
    for row in cable_rows:
        if row.get("sizeMm2") == size:
            return float(row["resistanceOhmPerKm"])
    return R_BY_SIZE.get(int(size))


def style_header(ws, row: int, cols: int) -> None:
    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(bold=True, color="FFFFFF")
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)


def autosize(ws, max_width: int = 28) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = min(max(len(str(c.value or "")) for c in col) + 2, max_width)
        ws.column_dimensions[letter].width = width


def main() -> None:
    db = json.loads(DATA.read_text(encoding="utf-8"))
    machines = {m["id"]: m for m in db["machines"]}
    copper = db["copperKFactors"]
    cable_cap = db["cableCapacity"]

    wb = Workbook()

    # Sheet 1 — v0 machines (used parameters)
    ws = wb.active
    ws.title = "v0 machines (used)"
    headers = [
        "Model (col D)",
        "Manufacturer",
        "Category",
        "I crank T (col T) — USED",
        "Alternator Z (col Z)",
        "Cable type AD",
        "Size AE mm²",
        "Continuous AG A",
        "Site temp AF °C",
        "K (lookup)",
        "R Ω/km (lookup)",
        "Q col T — NOT USED",
        "R inrush — NOT USED",
        "Power cutoff kW",
        "Cutoff V",
        "Efficiency %",
    ]
    ws.append(headers)
    style_header(ws, 1, len(headers))

    for mid in CLIENT_IDS:
        m = machines[mid]
        size = m.get("cableSizeMm2")
        ws.append(
            [
                mid,
                m.get("manufacturer"),
                m.get("category"),
                m.get("peakCrankingCurrentA"),
                m.get("alternatorContinuousA"),
                m.get("cableType"),
                size,
                m.get("cableContinuousA"),
                m.get("operatingTempC"),
                k_factor(str(m.get("cableType", "")), copper),
                resistance(size, cable_cap),
                m.get("peakCurrentCutoffA"),
                m.get("inrushCurrentA"),
                m.get("powerAtCutoffKw"),
                m.get("cutoffVoltageV"),
                m.get("efficiencyPercent"),
            ]
        )
    autosize(ws)

    # Sheet 2 — user inputs & constants
    ws2 = wb.create_sheet("Inputs & constants")
    ws2.append(["Type", "Parameter", "PDF line", "Value / notes"])
    style_header(ws2, 1, 4)
    rows = [
        ("User", "Safety factor %", "1", "25 or 50"),
        ("User", "Machine model", "2", "9 machines"),
        ("User", "Battery V during cranking", "3", "User entered"),
        ("User", "Operating temp °C", "4", "User entered"),
        ("Constant", "Min starter voltage V", "5", "16"),
        ("Constant", "Cranking time s", "11", "5"),
    ]
    for r in rows:
        ws2.append(list(r))
    autosize(ws2)

    # Sheet 3 — full fleet T vs Q eligibility
    ws3 = wb.create_sheet("Fleet T vs Q")
    h3 = [
        "Model",
        "In v0 UI",
        "T path OK",
        "Q path OK",
        "T (A)",
        "Q (A)",
        "R inrush (A)",
        "Alternator (A)",
        "Cable type",
        "Size mm²",
    ]
    ws3.append(h3)
    style_header(ws3, 1, len(h3))

    def gba_ok(m: dict, use_q: bool) -> bool:
        if use_q:
            q = m.get("peakCurrentCutoffA")
            try:
                qn = float(q)
                if qn <= 0:
                    return False
            except (TypeError, ValueError):
                return False
        else:
            try:
                if float(m.get("peakCrankingCurrentA", 0)) <= 0:
                    return False
            except (TypeError, ValueError):
                return False
        try:
            if float(m.get("alternatorContinuousA", 0)) <= 0:
                return False
            if float(m.get("cableSizeMm2", 0)) <= 0:
                return False
        except (TypeError, ValueError):
            return False
        return bool(m.get("cableType"))

    for m in sorted(db["machines"], key=lambda x: x["id"]):
        ws3.append(
            [
                m["id"],
                "Yes" if m["id"] in CLIENT_IDS else "No",
                "Yes" if gba_ok(m, False) else "No",
                "Yes" if gba_ok(m, True) else "No",
                m.get("peakCrankingCurrentA"),
                m.get("peakCurrentCutoffA"),
                m.get("inrushCurrentA"),
                m.get("alternatorContinuousA"),
                m.get("cableType"),
                m.get("cableSizeMm2"),
            ]
        )
    autosize(ws3)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
